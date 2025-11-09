import re
from copy import copy
from datetime import datetime
from typing import Any, Dict, List, Union

import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.header_footer import HeaderFooterItem
from openpyxl.worksheet.merge import MergedCellRange

from schemas.templater import TemplaterRequest
from utils.logger import GLOBAL_LOGGER
from utils.unitofwork import IUnitOfWork

from .base_templater import BaseTemplater


class TemplaterXlsx(BaseTemplater):
    """Класс для автоматической генерации справки по формату ".xlsx" ."""

    def __init__(
        self,
        uow: IUnitOfWork,
        enquiry_id: int,
        templater_params: TemplaterRequest,
    ):
        """Инициализирует генератор справки.

        Args:
            uow (IUnitOfWork): Unit of Work для доступа к базам данных.
            enquiry_id (int): ID справки для получения шаблонов.
            templater_params (TemplaterRequest): Параметры для фильтрации
                данных во внешних запросах.

        """
        super().__init__(
            uow, enquiry_id, templater_params, template_type="xlsx"
        )
        self.outputFile = "final.xlsx"

    def _load_template(self):
        self.book = openpyxl.load_workbook(self.template_file_path)
        self.sheet = self.book.active

    _SAMPLE_RESPONSES = {
        "5": [
            {"inn": 777777, "chinases": "МБОУ СОШ 13579"},
            {"inn": 777777, "chinases": "МБОУ СОШ 13579"},
            {"inn": 777777, "chinases": "МБОУ СОШ 13579"},
        ],
        "8": [{"inn": 777777}],
        "7": [{"inn": 777777}],
        "6": [
            {
                "n": 1,
                "pokaz": "Работы",
                "n2": 3,
                "pokaz34": "Ыбок",
                "y_2023": 55.66,
                "y_2024": 66.77,
                "y_2025": 100.25,
            },
            {
                "n": 2,
                "pokaz": "Услуги",
                "n2": 4,
                "pokaz34": "ЫЫЫБ",
                "y_2023": 55.66,
                "y_2024": 66.77,
                "y_2025": 100.25,
            },
            {
                "n": 3,
                "pokaz": "Работы",
                "n2": 77,
                "pokaz34": "РЫБЫ",
                "y_2023": 55.662,
                "y_2024": 66.737,
                "y_2025": 100.425,
            },
            {
                "n": 4,
                "pokaz": "РаботыS",
                "n2": 52,
                "pokaz34": "workЫ",
                "y_2023": 55.663,
                "y_2024": 66.727,
                "y_2025": 100.255,
            },
        ],
        "9": [
            {
                "n": 5,
                "pokaz": "PIPIPUPU",
                "y_2023": 255.66,
                "y_2024": 366.77,
                "y_2025": 4100.25,
            },
            {
                "n": 7,
                "pokaz": "PIPIPUPU",
                "y_2023": 255.66,
                "y_2024": 366.77,
                "y_2025": 4100.25,
            },
            {
                "n": 6,
                "pokaz": "PIPIPUPU",
                "y_2023": 255.662,
                "y_2024": 366.737,
                "y_2025": 4100.425,
            },
        ],
    }

    def _add_footer_info(self) -> None:
        """Добавляет нижний колонтитул с датой и временем генерации в лист.

        Функция добавляет информацию о времени формирования справки в правую
        часть нижнего колонтитула активного листа.

        Args:
            No arguments.

        Returns:
            None: Функция модифицирует объект `self.sheet` напрямую.
        """
        footer = self.sheet.oddFooter
        date_and_time = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        credentials_text = f"Справка сформирована\n{date_and_time}"

        footer.right.text = credentials_text
        footer.right.size = 9
        footer.right.font = "Times New Roman"

    async def _replace_single_placeholders(self) -> None:
        """Замещает одиночные плейсхолдеры в ячейках листа.

        Фунция итерируется по ячейкам листа и ищет плейсхолдеры, формата
        `{{query_id:response_index;column_name}}`. Найденные плейсхолдеры
        заменяются на соответствующие данные из `self.query_responses`.

        Args:
            No arguments.

        Returns:
            None: Функция модифицирует объект `self.sheet` напрямую.
        """
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    match = self.combined_pattern.search(cell.value)
                    if match:
                        updated_value = await self._placeholder_replacer(match)
                        cell.value = cell.value.replace(
                            match.group(0), updated_value
                        )

    def _process_dynamic_tables(self) -> None:
        """Заполняет лист данными.

        Вызывает функцию для сбора данных о форматировании листа и произ-
        водит замену однострочных плейсхолдеров. Затем подготавливает лист,
        сдвигая диапазоны ячеек для вставки данных,после заполняет шаблонные
        строки данными из ответов на запросы.

        Args:
            No arguments.

        Returns:
            None: Функция модифицирует объект `self.sheet` напрямую.
        """
        row_layouts = self.collect_layout()
        current_row_index = 1
        while current_row_index <= self.sheet.max_row:
            row = self.sheet[current_row_index]
            current_row_index += 1
            for cell in row:
                if cell.column == len(row) and row_layouts.get(cell.row):
                    if (
                        row_layouts[cell.row].get("moveRange")
                        or row_layouts[cell.row].get("moveRange") == 0
                    ):
                        if row_layouts[cell.row].get("moveRange") != 0:
                            range_to_move = f"{get_column_letter(1)}{
                                (cell.row + 1)
                            }:{get_column_letter(self.sheet.max_column)}{
                                self.sheet.max_row
                            }"
                            rows_to_insert = row_layouts[cell.row]["moveRange"]
                            query_ids = [
                                int(key)
                                for key in row_layouts[cell.row]
                                if key.isdigit()
                                and int(key) in self.query_responses
                            ]

                            self.sheet.move_range(
                                range_to_move, rows_to_insert
                            )
                            for merged in self.sheet.merged_cells.ranges:
                                if merged.min_row >= (cell.row + 1):
                                    merged.min_row += rows_to_insert
                                    merged.max_row += rows_to_insert
                            if not isinstance(query_ids, list):
                                query_ids = [query_ids]

                            for query_id in query_ids:
                                data_block = self.query_responses[query_id]
                                self.process_data_block(
                                    cell,
                                    row_layouts,
                                    data_block,
                                    query_id,
                                )
                        else:
                            query_ids = list()
                            for key in row_layouts[cell.row]:
                                if (
                                    key.isdigit()
                                    and int(key) in self.query_responses
                                ):
                                    query_ids.append(int(key))
                            for query_id in query_ids:
                                data_block = self.query_responses[query_id]
                                self.process_data_block(
                                    cell,
                                    row_layouts,
                                    data_block,
                                    query_id,
                                )
                    else:
                        self.sheet.row_dimensions[
                            cell.row
                        ].height = row_layouts[cell.row]["layout"]["height"]

    def _save_and_cleanup(self) -> str:
        """Сохраняет итоговый файл."""
        self.book.save(self.outputFile)
        GLOBAL_LOGGER.info(f"File saved to: {self.outputFile}")
        return self.outputFile

    def process_data_block(
        self,
        cell: openpyxl.cell.cell.Cell,
        row_layouts: dict,
        data_block: List[Dict[str, Any]],
        query_id: int,
    ) -> None:
        """Обрабатывает блок таблицы данными.

        Функция итерируется по списку `data_block` и для каждого элемента
        создает новую строку в таблице, за исключением первой(шаблонной),
        начиная со строки определенной ячейкой `cell`. Стили и структура новой
        строки копируются из шаблонной, описанной в `row_layouts`.

        Args:
            cell (openpyxl.cell.cell.Cell): Ячейка, определяющая начальную
            строку для вставки блока данных.
            row_layouts (dict): Словарь с информацией о форматировании
            и ключах для `self.query_responses` для шаблонных строк.
            data_block (list): Список словарей с данными для заполнения. Каждый
            внутренний словарь соответствует одной строке.
            query_id (int): Id запроса(`query_id`) используемый для
            фильтрации плейсхолдеров в шаблонной строке.

        Returns:
            None: Функция модифицирует объект `self.sheet` напрямую.
        """
        for data_index, data_item in enumerate(data_block):
            target_row_index = cell.row + data_index
            self.sheet.row_dimensions[target_row_index].height = row_layouts[
                cell.row
            ]["layout"]["height"]

            for column_index, cell_properties in enumerate(
                row_layouts[cell.row]["layout"]["cells"], 1
            ):
                new_cell = self.sheet.cell(
                    row=target_row_index, column=column_index
                )
                new_cell.font = copy(cell_properties["font"])
                new_cell.border = copy(cell_properties["border"])
                new_cell.fill = copy(cell_properties["fill"])
                new_cell.number_format = cell_properties["number_format"]
                new_cell.alignment = copy(cell_properties["alignment"])

                placeholder_text = cell_properties["value"]

                if (
                    placeholder_text
                    and isinstance(placeholder_text, str)
                    and not isinstance(new_cell, MergedCell)
                ):
                    placeholder_match = re.findall(
                        r"\{\{(\d+);(\w+)\}\}", placeholder_text
                    )
                    if (
                        placeholder_match
                        and int(placeholder_match[0][0]) == query_id
                    ):
                        value_to_insert = data_item.get(
                            placeholder_match[0][1], ""
                        )
                        if isinstance(value_to_insert, str):
                            placeholder_text = placeholder_text.replace(
                                placeholder_text, value_to_insert
                            )
                        else:
                            placeholder_text = placeholder_text.replace(
                                placeholder_text, str(value_to_insert)
                            )
                            try:
                                placeholder_text = int(placeholder_text)
                            except ValueError:
                                try:
                                    placeholder_text = float(placeholder_text)
                                except ValueError:
                                    pass

                        new_cell.value = placeholder_text
                    elif not placeholder_match and not isinstance(
                        new_cell, MergedCell
                    ):
                        new_cell.value = placeholder_text

            for merged_range in row_layouts[cell.row]["layout"]["merges"]:
                row_offset = target_row_index - merged_range.min_row
                new_start_row = merged_range.min_row + row_offset
                new_end_row = merged_range.max_row + row_offset
                self.sheet.merge_cells(
                    start_row=new_start_row,
                    start_column=merged_range.min_col,
                    end_row=new_end_row,
                    end_column=merged_range.max_col,
                )

    def capture_row_layout(
        self, row_index: int
    ) -> Dict[str, Union[int, list]]:
        """Собирает форматирование ряда.

        Args:
            row_index (int): Ряд для сбора форматирования.

        Returns:
            dict: Функция возвращает словарь содержащий информацию о
            форматировании ряда.

        """
        row_properties = {
            "height": self.sheet.row_dimensions[row_index].height,
            "cells": [],
            "merges": [],
        }

        # Получаем свойства клетки
        for cell in self.sheet[row_index]:
            row_properties["cells"].append(
                {
                    "value": cell.value,
                    "font": copy(cell.font),
                    "border": copy(cell.border),
                    "fill": copy(cell.fill),
                    "number_format": cell.number_format,
                    "alignment": copy(cell.alignment),
                }
            )

        # Получаем сведения о всех смерженных клетках в ряду
        for merged_range in self.sheet.merged_cells.ranges:
            if merged_range.min_row <= row_index <= merged_range.max_row:
                row_properties["merges"].append(merged_range)

        return row_properties

    def collect_layout(self) -> dict:
        """Собирает форматирвание листа.

        Функция итерируется по листу и собирает информацию о высоте, стилях
        и прочих параметрах ячеек. Сбор форматирования собирается вокруг
        плейсхолдеров блоков - `{{query_id;column_name}}`.

        Args:
            No arguments.

        Returns:
           dict: Словарь с информацией о форматировании строк,
          индексированный по номеру строки в итоговом документе.
        """
        block_pattern = r"\{\{(\d+);(\w+)\}\}"
        formatter_pattern = r"\{\{([a-zA-Z_][\w;]*);(\d+;[a-zA-Z_]\w*)\}\}"
        row_offset = 0
        row_layouts: dict[
            int,
            dict[
                str,
                dict[str, int | list[MergedCellRange | dict[str, str | Font]]]
                | list[str]
                | int,
            ],
        ] = {}
        rows_to_delete = list()
        all_block_matches = list()
        last_placeholder_row_index = 0
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    block_match = re.findall(block_pattern, cell.value)
                    # formatter_match = re.findall(formatter_pattern, cell.value)

                    # if formatter_match:
                    #     current_result = 0
                    #     formatter_operations =
                    #     formatter_match[0][0].split(";")
                    #     response_idx, response_key = formatter_match[0][
                    #         1
                    #     ].split(";")
                    #     for operation_name in reversed(formatter_operations):
                    #         operation_name = operation_name.lower()
                    #         match operation_name:
                    #             case "sum":
                    #                 for values in self.query_responses[
                    #                     int(response_idx)
                    #                 ]:
                    #                     current_result +=
                    #     values[response_key]
                    #             case "round":
                    #                 if not current_result:
                    #                     for values in self.query_responses[
                    #                         int(response_idx)
                    #                     ]:
                    #                         current_result = round(
                    #                             values[response_key]
                    #                         )
                    #                 else:
                    #                     current_result =
                    #      round(current_result)
                    #             case "SEPNUM":
                    #                 pass
                    #     cell.value = current_result

                    if block_match:
                        all_block_matches.append(block_match)
                        if (
                            int(block_match[0][0]) not in self.query_responses
                            and (cell.row, cell.row + row_offset)
                            not in rows_to_delete
                        ):
                            rows_to_delete.append(
                                (cell.row, cell.row + row_offset)
                            )
                        if last_placeholder_row_index and (
                            cell.row - last_placeholder_row_index != 1
                        ):
                            for header_row in self.sheet.iter_rows(
                                min_row=last_placeholder_row_index + 1,
                                max_row=cell.row - 1,
                            ):
                                if not row_layouts.get(
                                    header_row[0].row + row_offset
                                ):
                                    row_layouts[
                                        header_row[0].row + row_offset
                                    ] = {
                                        "layout": self.capture_row_layout(
                                            header_row[0].row
                                        )
                                    }
                        for query_id_str, placeholder_key in block_match:
                            if not row_layouts.get(cell.row + row_offset):
                                row_layouts[cell.row + row_offset] = dict()
                                row_layouts[cell.row + row_offset][
                                    "layout"
                                ] = self.capture_row_layout(cell.row)
                            if not row_layouts[cell.row + row_offset].get(
                                query_id_str
                            ):
                                row_layouts[cell.row + row_offset][
                                    query_id_str
                                ] = list()
                            layout_value = row_layouts[cell.row + row_offset][
                                query_id_str
                            ]
                            if isinstance(
                                layout_value,
                                list,
                            ):
                                layout_value.append(placeholder_key)

                if cell.column == len(row) and row_layouts.get(
                    cell.row + row_offset
                ):
                    last_placeholder_row_index = cell.row
                    requests_used = len(
                        [
                            key
                            for key in row_layouts[cell.row + row_offset]
                            if key.isdigit()
                            and int(key) in self.query_responses
                        ]
                    )
                    rows_to_insert = 0
                    if requests_used > 1:
                        max_rows = 0
                        for key in row_layouts[cell.row + row_offset]:
                            if (
                                key.isdigit()
                                and int(key) in self.query_responses
                            ):
                                if max_rows:
                                    if (
                                        len(self.query_responses[int(key)])
                                        > max_rows
                                    ):
                                        max_rows = len(
                                            self.query_responses[int(key)]
                                        )
                                else:
                                    max_rows = len(
                                        self.query_responses[int(key)]
                                    )
                        rows_to_insert = max_rows - 1
                        row_offset += rows_to_insert
                    else:
                        for key in row_layouts[cell.row + row_offset]:
                            if (
                                key.isdigit()
                                and int(key) in self.query_responses
                            ):
                                rows_to_insert = (
                                    len(self.query_responses[int(key)]) - 1
                                )
                                row_offset += rows_to_insert
                    row_layouts[list(row_layouts.keys())[-1]]["moveRange"] = (
                        rows_to_insert
                    )

        for template_row_num, final_row_num in reversed(rows_to_delete):
            for merged in list(self.sheet.merged_cells.ranges):
                if merged.min_row > template_row_num:
                    merged.min_row -= 1
                    merged.max_row -= 1
                elif merged.min_row <= template_row_num <= merged.max_row:
                    if merged.min_row == merged.max_row:
                        self.sheet.merged_cells.remove(merged)
                    else:
                        merged.max_row -= 1
            if row_layouts.get(final_row_num):
                del row_layouts[final_row_num]
                rows_to_update = [
                    row for row in row_layouts if row > final_row_num
                ]
                for old_row in rows_to_update:
                    new_row = old_row - 1
                    row_layouts[new_row] = row_layouts.pop(old_row)
                self.sheet.delete_rows(template_row_num, 1)

        return row_layouts

    def unescape_newline(self):
        """Monkey patch для символа новой строки в openpyxl."""
        original_value = self.__str_orig__()
        return original_value.replace("_x000a_", "\n")

    if not hasattr(HeaderFooterItem, "__str_orig__"):
        HeaderFooterItem.__str_orig__ = HeaderFooterItem.__str__
        HeaderFooterItem.__str__ = unescape_newline
